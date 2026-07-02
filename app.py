import streamlit as st
import google.generativeai as genai
import pandas as pd
import os
import io
import json
from PIL import Image

# ==============================================================================
# 1. CẤU HÌNH TRANG & GIAO DIỆN (THEME & STYLING)
# ==============================================================================
st.set_page_config(
    page_title="AI Invoice Extractor - Trích Xuất Hóa Đơn Tự Động",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS để giao diện trông chuyên nghiệp và hiện đại hơn
st.markdown("""
<style>
    /* Chỉnh sửa tổng thể */
    .main-title {
        font-size: 2.6rem;
        color: #1E3A8A;
        font-weight: 800;
        margin-bottom: 0.5rem;
        text-align: center;
    }
    .subtitle {
        font-size: 1.1rem;
        color: #4B5563;
        margin-bottom: 2rem;
        text-align: center;
    }
    /* Tùy chỉnh các khối thông báo */
    .stAlert {
        border-radius: 8px;
    }
    /* Chỉnh sửa bảng preview */
    .dataframe {
        border-radius: 8px;
        overflow: hidden;
    }
    /* Giao diện form đăng nhập */
    .login-box {
        padding: 2.5rem;
        border-radius: 10px;
        background-color: #FFFFFF;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        border: 1px solid #E5E7EB;
    }
</style>
""", unsafe_allow_html=True)

# ==============================================================================
# 2. XÁC THỰC NGƯỜI DÙNG (AUTHENTICATION SYSTEM)
# ==============================================================================
# Khởi tạo trạng thái phiên đăng nhập nếu chưa tồn tại
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

def check_login(email, password):
    """
    Kiểm tra thông tin đăng nhập từ cấu hình st.secrets["credentials"].
    Tuyệt đối không hardcode tài khoản mật khẩu vào code.
    """
    if "credentials" in st.secrets:
        credentials = st.secrets["credentials"]
        # So sánh email và mật khẩu tương ứng trong danh sách
        if email in credentials and str(credentials[email]) == password:
            return True
    return False

# ==============================================================================
# 3. XÁC THỰC API KEY BẢO MẬT
# ==============================================================================
def get_api_key():
    """
    Lấy API Key từ Streamlit Secrets hoặc Biến môi trường hệ thống.
    Đảm bảo an toàn bảo mật, không lộ Key lên mã nguồn.
    """
    # 1. Tìm trong Streamlit Secrets (Dùng khi deploy lên Streamlit Cloud)
    if "GEMINI_API_KEY" in st.secrets:
        return st.secrets["GEMINI_API_KEY"]
    
    # 2. Tìm trong biến môi trường hệ thống
    api_key = os.environ.get("GEMINI_API_KEY")
    if api_key:
        return api_key
        
    return None

api_key = get_api_key()

# ==============================================================================
# 4. ĐỊNH NGHĨA SCHEMA CHO TRÍCH XUẤT DỮ LIỆU CÓ CẤU TRÚC (GEMINI SCHEMA DICT)
# ==============================================================================
invoice_schema = {
    "type": "OBJECT",
    "properties": {
        "invoice_number": {
            "type": "STRING",
            "description": "Số hóa đơn (Invoice number/No.). Nếu không tìm thấy, trả về null."
        },
        "invoice_date": {
            "type": "STRING",
            "description": "Ngày lập hóa đơn (Invoice date). Định dạng chuỗi chuẩn DD/MM/YYYY. Nếu không tìm thấy, trả về null."
        },
        "seller_tax_code": {
            "type": "STRING",
            "description": "Mã số thuế của bên bán / đơn vị cung cấp (Seller's Tax Code/MST). Nếu không tìm thấy, trả về null."
        },
        "seller_name": {
            "type": "STRING",
            "description": "Tên công ty / đơn vị bán hàng (Seller's Name). Nếu không tìm thấy, trả về null."
        },
        "total_before_tax": {
            "type": "NUMBER",
            "description": "Tổng tiền hàng chưa thuế / giá trị trước thuế (Total before tax/Subtotal). Trả về dạng số thực (float). Nếu không tìm thấy, trả về null."
        },
        "tax_amount": {
            "type": "NUMBER",
            "description": "Tiền thuế GTGT / VAT (Tax amount). Trả về dạng số thực (float). Nếu không tìm thấy, trả về null."
        },
        "total_amount": {
            "type": "NUMBER",
            "description": "Tổng cộng tiền thanh toán đã bao gồm thuế (Total payment amount/Total after tax). Trả về dạng số thực (float). Nếu không tìm thấy, trả về null."
        }
    },
    "required": [
        "invoice_number", "invoice_date", "seller_tax_code", 
        "seller_name", "total_before_tax", "tax_amount", "total_amount"
    ]
}

# ==============================================================================
# 5. LOGIC XỬ LÝ BACKEND (GỌI GEMINI API & CƠ CHẾ SELF-CHECKING)
# ==============================================================================
def extract_invoice_data(api_key, file_bytes, mime_type, file_name, status_placeholder=None):
    """
    Gửi file nhị phân sang Gemini API để phân tích cấu trúc và trích xuất dữ liệu.
    Hỗ trợ cơ chế Auto-Retry tự động thử lại trực quan khi gặp lỗi Rate Limit (HTTP 429 / ResourceExhausted).
    """
    import time
    import re
    
    max_retries = 5
    base_delay = 10.0
    
    for attempt in range(max_retries):
        try:
            # Cấu hình API Key cho thư viện
            genai.configure(api_key=api_key)
            
            # Khởi tạo mô hình Gemini 2.5 Flash
            model = genai.GenerativeModel('gemini-1.5-flash')  # Quota miễn phí: 1,500 req/ngày (cao hơn 6x so với gemini-2.5-flash)
            
            # Đóng gói file thành format đầu vào của Gemini Multimodal API
            file_part = {
                "mime_type": mime_type,
                "data": file_bytes
            }
            
            prompt = (
                "Bạn là một trợ lý AI chuyên nghiệp về kế toán và xử lý tài liệu tại Việt Nam.\n"
                "Hãy đọc và phân tích kỹ tài liệu hóa đơn được cung cấp (dạng PDF hoặc hình ảnh).\n"
                "Trích xuất chính xác các thông tin cần thiết theo đúng cấu trúc yêu cầu.\n"
                "Lưu ý quan trọng:\n"
                "1. Đối với các trường số tiền (total_before_tax, tax_amount, total_amount), "
                "hãy chuyển đổi về dạng số thập phân thuần túy (float), loại bỏ mọi dấu phân cách hàng nghìn (ví dụ: dấu chấm hay phẩy) và đơn vị tiền tệ VND.\n"
                "Ví dụ: '1.500.000 đ' -> 1500000.0\n"
                "2. Đảm bảo mã số thuế người bán (seller_tax_code) là một chuỗi số viết liền, bao gồm cả chi nhánh phụ nếu có (ví dụ: '0101234567-001').\n"
                "3. Nếu thông tin nào không tồn tại trong hóa đơn, hãy trả về null."
            )
            
            # Gọi API với cấu hình Structured Output thông qua response_schema dạng dict
            response = model.generate_content(
                [prompt, file_part],
                generation_config=genai.GenerationConfig(
                    response_mime_type="application/json",
                    response_schema=invoice_schema,
                    temperature=0.1  # Giảm tính ngẫu nhiên để tăng độ chính xác trích xuất
                )
            )
            
            # Chuyển đổi chuỗi JSON phản hồi thành Python dict
            result_json = json.loads(response.text)
            result_json["file_name"] = file_name
            result_json["status"] = "Thành công"
            result_json["error_message"] = ""
            
            return result_json

        except Exception as e:
            error_str = str(e)
            # Kiểm tra xem có phải lỗi Rate Limit (429, Quota exceeded, ResourceExhausted) không
            is_rate_limit = any(keyword in error_str.lower() for keyword in ["429", "quota", "resourceexhausted", "limit"])
            
            if is_rate_limit and attempt < max_retries - 1:
                # Trích xuất số giây cần chờ từ lỗi nếu có (ví dụ: "Please retry in 48.04s")
                wait_time = base_delay
                match = re.search(r"retry in (\d+(?:\.\d+)?)s", error_str)
                if match:
                    wait_time = float(match.group(1)) + 2.0  # Cộng thêm 2 giây an toàn
                
                # Cập nhật cảnh báo lên giao diện để người dùng theo dõi
                if status_placeholder:
                    status_placeholder.warning(
                        f"⚠️ File `{file_name}` bị quá tải API (Lỗi 429). "
                        f"Hệ thống tự động chờ {wait_time:.1f} giây để reset băng thông và thử lại (Lần {attempt + 1}/{max_retries})..."
                    )
                
                # Tạm dừng và quay lại vòng lặp để thử lại
                time.sleep(wait_time)
                continue
            
            # Nếu là lỗi khác hoặc đã hết số lần thử lại
            return {
                "file_name": file_name,
                "invoice_number": None,
                "invoice_date": None,
                "seller_tax_code": None,
                "seller_name": None,
                "total_before_tax": None,
                "tax_amount": None,
                "total_amount": None,
                "status": "Thất bại",
                "error_message": f"Lỗi xử lý: {error_str}"
            }

# ==============================================================================
# 6. GIAO DIỆN PHẦN MỀM (FRONTEND)
# ==============================================================================

# KỊCH BẢN 1: NẾU NGƯỜI DÙNG CHƯA ĐĂNG NHẬP -> HIỂN THỊ MÀN HÌNH LOGIN
if not st.session_state["authenticated"]:
    st.markdown("<h1 class='main-title'>🧾 AI Invoice Extractor</h1>", unsafe_allow_html=True)
    st.markdown("<p class='subtitle'>Hệ thống Trích xuất Hóa đơn Tự động bằng Trí tuệ nhân tạo</p>", unsafe_allow_html=True)
    
    # Tạo Form đăng nhập căn giữa
    col1, col2, col3 = st.columns([1, 1.8, 1])
    with col2:
        st.markdown("<div class='login-box'>", unsafe_allow_html=True)
        st.markdown("<h3 style='text-align: center; color: #1E3A8A; margin-top: 0;'>🔐 ĐĂNG NHẬP HỆ THỐNG</h3>", unsafe_allow_html=True)
        
        with st.form("login_form"):
            email_input = st.text_input("📧 Email tài khoản", placeholder="nhap-email@example.com")
            password_input = st.text_input("🔑 Mật khẩu", type="password", placeholder="••••••••")
            login_submit = st.form_submit_button("Đăng nhập", use_container_width=True)
            
            if login_submit:
                # Kiểm tra cấu hình credentials trước
                if "credentials" not in st.secrets:
                    st.error("⚠️ Hệ thống chưa được cấu hình danh sách tài khoản (credentials) trong Secrets!")
                elif check_login(email_input, password_input):
                    st.session_state["authenticated"] = True
                    st.success("Đăng nhập thành công! Đang tải hệ thống...")
                    st.rerun()
                else:
                    st.error("❌ Tài khoản không hợp lệ hoặc sai mật khẩu")
        st.markdown("</div>", unsafe_allow_html=True)

# KỊCH BẢN 2: NẾU ĐÃ ĐĂNG NHẬP THÀNH CÔNG -> HIỂN THỊ TOÀN BỘ CHỨC NĂNG CHÍNH
else:
    # Sidebar hướng dẫn cấu hình & thông tin dự án
    with st.sidebar:
        st.image("https://img.icons8.com/clouds/200/000000/invoice.png", width=120)
        st.markdown("### 📑 Hướng Dẫn Sử Dụng")
        st.markdown(
            """
            1. **Tải lên hóa đơn**: Nhấn nút browse hoặc kéo thả các file hóa đơn dạng **PDF** hoặc **ảnh (PNG, JPG, JPEG)** vào khung tải lên.
            2. **Trích xuất dữ liệu**: Nhấn nút **Bắt đầu trích xuất**. Hệ thống sẽ gọi Gemini AI xử lý song song từng file.
            3. **Xem kết quả**: Kiểm tra bảng dữ liệu tổng hợp trực tiếp trên web.
            4. **Xuất Excel**: Tải file Excel chứa toàn bộ thông tin đã trích xuất về máy.
            """
        )
        st.markdown("---")
        st.markdown("### 🔒 Bảo Mật Thông Tin")
        st.info(
            "API Key được quản lý an toàn qua biến môi trường (Streamlit Secrets). "
            "Dữ liệu hóa đơn của bạn chỉ được xử lý qua API và không được lưu trữ lại trên máy chủ."
        )
        st.markdown("---")
        # Nút đăng xuất ở sidebar
        if st.button("🚪 Đăng xuất tài khoản", use_container_width=True, type="secondary"):
            st.session_state["authenticated"] = False
            st.rerun()
        st.markdown("---")
        st.markdown("⚡ *Powered by Gemini 2.5 Flash*")

    # Phần tiêu đề chính của trang chức năng
    st.markdown("<h1 class='main-title'>🧾 AI Invoice Extractor</h1>", unsafe_allow_html=True)
    st.markdown("<p class='subtitle'>Trích xuất hóa đơn PDF & Ảnh sang Excel tự động bằng Trí tuệ nhân tạo Gemini</p>", unsafe_allow_html=True)

    # Kiểm tra trạng thái cấu hình API Key
    if not api_key:
        st.warning("⚠️ **Hệ thống chưa được cấu hình API Key!**")
        st.markdown(
            """
            **Hướng dẫn thiết lập nhanh:**
            - **Khi chạy local**: Tạo file `.streamlit/secrets.toml` trong thư mục gốc của dự án và dán nội dung sau:
              ```toml
              GEMINI_API_KEY = "Nhập_API_Key_Gemini_Của_Bạn_Vào_Đây"
              ```
            - **Khi deploy lên Streamlit Cloud**: Vào cài đặt ứng dụng (App settings) -> mục **Secrets** và dán:
              ```toml
              GEMINI_API_KEY = "Nhập_API_Key_Gemini_Của_Bạn_Vào_Đây"
              ```
            """
        )
        # Cho phép người dùng nhập key tạm thời trực tiếp trên giao diện để dùng thử nếu muốn
        temp_key = st.text_input("Hoặc nhập tạm thời API Key của bạn tại đây để trải nghiệm ngay:", type="password")
        if temp_key:
            api_key = temp_key
            st.success("Đã ghi nhận API Key tạm thời!")

    # Nếu đã có API Key, tiếp tục hiển thị ứng dụng
    if api_key:
        # 1. Khung upload file hóa đơn
        uploaded_files = st.file_uploader(
            "Tải lên các file hóa đơn (Hỗ trợ PDF, PNG, JPG, JPEG) - Có thể chọn nhiều file cùng lúc",
            type=["pdf", "png", "jpg", "jpeg"],
            accept_multiple_files=True
        )
        
        if uploaded_files:
            st.write(f"📂 Đã chọn **{len(uploaded_files)}** file để xử lý.")
            
            # Nút bấm bắt đầu trích xuất
            if st.button("🚀 Bắt đầu trích xuất dữ liệu", use_container_width=True):
                results = []
                
                # Khởi tạo thanh tiến trình và trạng thái
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                # Vòng lặp xử lý từng file
                for idx, uploaded_file in enumerate(uploaded_files):
                    file_name = uploaded_file.name
                    status_text.markdown(f"⏳ Đang xử lý file ({idx + 1}/{len(uploaded_files)}): `{file_name}`...")
                    
                    # Đọc dữ liệu nhị phân của file
                    file_bytes = uploaded_file.read()
                    
                    # Xác định Mime Type của file dựa trên đuôi mở rộng
                    file_extension = file_name.split('.')[-1].lower()
                    if file_extension == 'pdf':
                        mime_type = 'application/pdf'
                    elif file_extension in ['png', 'jpg', 'jpeg']:
                        mime_type = f'image/{file_extension if file_extension != "jpg" else "jpeg"}'
                    else:
                        mime_type = 'application/octet-stream'
                    
                    # Tự kiểm tra định dạng và dữ liệu rỗng (Self-Checking)
                    if len(file_bytes) == 0:
                        results.append({
                            "file_name": file_name,
                            "invoice_number": None,
                            "invoice_date": None,
                            "seller_tax_code": None,
                            "seller_name": None,
                            "total_before_tax": None,
                            "tax_amount": None,
                            "total_amount": None,
                            "status": "Thất bại",
                            "error_message": "Lỗi: File trống (dung lượng 0 bytes)."
                        })
                    else:
                        # Nếu không phải file đầu tiên, thêm khoảng nghỉ ngắn 6 giây
                        # để tránh lỗi Rate Limit (Too Many Requests - HTTP 429) của gói Gemini Free Tier (15 RPM)
                        if idx > 0:
                            status_text.markdown(f"⏳ Đang chờ giãn cách 6 giây để bảo vệ API... (File {idx + 1}/{len(uploaded_files)}: `{file_name}`)")
                            import time
                            time.sleep(6.0)
                            status_text.markdown(f"⏳ Đang xử lý file ({idx + 1}/{len(uploaded_files)}): `{file_name}`...")
                        
                        # Gọi hàm trích xuất dữ liệu từ Backend (truyền thêm status_text để hiển thị cảnh báo thử lại nếu quá tải)
                        extracted_data = extract_invoice_data(api_key, file_bytes, mime_type, file_name, status_placeholder=status_text)
                        results.append(extracted_data)
                    
                    # Cập nhật thanh tiến trình
                    progress = (idx + 1) / len(uploaded_files)
                    progress_bar.progress(progress)
                
                status_text.success("🎉 Đã hoàn thành xử lý toàn bộ các hóa đơn!")
                
                # 2. Xử lý hiển thị kết quả bằng Pandas
                df = pd.DataFrame(results)
                
                # Sắp xếp lại thứ tự cột cho hợp lý
                cols_order = [
                    "file_name", "invoice_number", "invoice_date", 
                    "seller_tax_code", "seller_name", "total_before_tax", 
                    "tax_amount", "total_amount", "status", "error_message"
                ]
                df = df[cols_order]
                
                # Chèn thêm cột số thứ tự (STT) bắt đầu từ 1 vào đầu bảng
                df.insert(0, "STT", range(1, len(df) + 1))
                
                # Việt hóa tiêu đề các cột để hiển thị trên bảng
                df_display = df.rename(columns={
                    "STT": "STT",
                    "file_name": "Tên File",
                    "invoice_number": "Số Hóa Đơn",
                    "invoice_date": "Ngày Lập",
                    "seller_tax_code": "Mã Số Thuế Bán",
                    "seller_name": "Đơn Vị Bán",
                    "total_before_tax": "Tiền Trước Thuế",
                    "tax_amount": "Tiền Thuế GTGT",
                    "total_amount": "Tổng Thanh Toán",
                    "status": "Trạng Thái",
                    "error_message": "Chi Tiết Lỗi"
                })
                
                # Hiển thị thống kê tổng quan
                success_count = sum(1 for r in results if r["status"] == "Thành công")
                fail_count = len(results) - success_count
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Tổng số file", len(results))
                with col2:
                    st.metric("Thành công ✅", success_count)
                with col3:
                    st.metric("Thất bại ❌", fail_count)
                
                # Hiển thị bảng kết quả Preview (ẩn cột index mặc định của Pandas)
                st.markdown("### 📊 Xem trước kết quả trích xuất")
                st.dataframe(df_display, use_container_width=True, hide_index=True)
                
                # 3. Tạo file Excel xuất ra bằng thư viện openpyxl qua Pandas
                # Lưu file Excel vào bộ nhớ đệm (BytesIO) để phục vụ việc download trực tiếp
                excel_buffer = io.BytesIO()
                from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    # Đẩy dataframe hiển thị (đã Việt hóa tiêu đề) vào sheet Excel
                    df_display.to_excel(writer, index=False, sheet_name="HoaDonTricXuat")
                    
                    workbook = writer.book
                    worksheet = writer.sheets["HoaDonTricXuat"]
                    
                    # Định nghĩa các kiểu định dạng (border, màu nền, font)
                    thin_border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )
                    
                    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')  # Màu xanh Navy đậm
                    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')  # Chữ trắng, in đậm
                    data_font = Font(name='Segoe UI', size=10)
                    
                    # Áp dụng định dạng (border, font, căn lề, format số) cho từng ô
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), start=1):
                        for col_idx, cell in enumerate(row, start=1):
                            # Áp dụng kẻ khung viền cho tất cả các ô
                            cell.border = thin_border
                            
                            if row_idx == 1:
                                # Định dạng cho hàng tiêu đề
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                # Định dạng cho hàng dữ liệu
                                cell.font = data_font
                                
                                # Căn lề số tiền sang bên phải (các cột: Tiền Trước Thuế, Tiền Thuế GTGT, Tổng Thanh Toán - nay là cột 7, 8, 9)
                                if col_idx in [7, 8, 9]:
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                                    # Định dạng hiển thị số tiền có dấu phân cách hàng nghìn (ví dụ: 1,500,000)
                                    if isinstance(cell.value, (int, float)):
                                        cell.number_format = '#,##0'
                                # Căn lề giữa cho các cột ngắn (STT, Số Hóa Đơn, Ngày Lập, MST Bán, Trạng Thái - nay là cột 1, 3, 4, 5, 10)
                                elif col_idx in [1, 3, 4, 5, 10]:
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                # Căn lề trái cho các cột văn bản dài (Tên File, Đơn Vị Bán, Chi Tiết Lỗi - nay là cột 2, 6, 11)
                                else:
                                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Tự động điều chỉnh độ rộng cột dựa trên nội dung đã định dạng
                    for col in worksheet.columns:
                        max_len = 0
                        for cell in col:
                            val_str = str(cell.value or '')
                            # Nếu là số tiền, tính độ dài dựa trên chuỗi đã format có dấu phẩy phân cách nghìn
                            if cell.number_format == '#,##0' and isinstance(cell.value, (int, float)):
                                val_str = f"{cell.value:,.0f}"
                            max_len = max(max_len, len(val_str))
                        col_letter = col[0].column_letter
                        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
                excel_data = excel_buffer.getvalue()
                
                # Nút tải xuống Excel
                st.download_button(
                    label="📥 Tải xuống kết quả Excel (.xlsx)",
                    data=excel_data,
                    file_name="ket_qua_trich_xuat_hoa_don.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # Hiển thị báo cáo lỗi chi tiết nếu có file xử lý thất bại
                if fail_count > 0:
                    st.markdown("### ⚠️ Nhật ký lỗi chi tiết")
                    for r in results:
                        if r["status"] == "Thất bại":
                            st.error(f"**File:** `{r['file_name']}` | **Lỗi:** {r['error_message']}")
